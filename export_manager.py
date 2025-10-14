"""
エクスポート管理モジュール（拡張版）
結果のエクスポート機能を管理 - PDF・Excel出力対応
"""

from tkinter import filedialog, messagebox, ttk
import tkinter as tk
from datetime import datetime
from security_manager import SecurityManager

# PDF出力用ライブラリ
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    PDF_AVAILABLE = True
except ImportError as e:
    print(f"PDF library import error: {e}")
    PDF_AVAILABLE = False
except Exception as e:
    print(f"PDF library initialization error: {e}")
    PDF_AVAILABLE = False

# Excel出力用ライブラリ
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import BarChart, Reference
    EXCEL_AVAILABLE = True
except ImportError as e:
    print(f"Excel library import error: {e}")
    EXCEL_AVAILABLE = False
except Exception as e:
    print(f"Excel library initialization error: {e}")
    EXCEL_AVAILABLE = False


class ExportManager:
    """エクスポート管理クラス（拡張版）"""
    
    def __init__(self, app):
        self.app = app
        self.security_manager = SecurityManager()
    
    def export_results(self):
        """結果のエクスポート（従来のテキスト出力）"""
        if not hasattr(self.app.controller, 'last_db_data') or not self.app.controller.last_db_data: 
            messagebox.showinfo("エクスポート不可", "先に計算を実行してください。")
            return
            
        # 結果テキストの生成
        texts = self.app.controller.ui_manager.generate_result_texts(
            self.app.controller.last_db_data, 
            self.app.controller.last_stats_results, 
            self.app.controller.last_inputs
        )
        sample_size_disp = self.app.controller.ui_manager.format_int(self.app.controller.last_stats_results['sample_size'])
        
        content = f"""AI SQC Sampler - 計算結果（AQL/LTPD設計）
計算日時: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
{'=' * 60}

品番: {self.app.controller.last_inputs['product_number']}
ロットサイズ: {self.app.controller.ui_manager.format_int(self.app.controller.last_inputs['lot_size'])}個
不具合率: {self.app.controller.last_db_data['defect_rate']:.2f}%
検査水準: {self.app.controller.last_stats_results['level_text']}
サンプルサイズ: {sample_size_disp} 個

【AQL/LTPD設計パラメータ】
AQL（合格品質水準）: {self.app.controller.last_stats_results.get('aql', self.app.controller.last_inputs.get('aql', 0.25)):.3f}%
LTPD（不合格品質水準）: {self.app.controller.last_stats_results.get('ltpd', self.app.controller.last_inputs.get('ltpd', 1.0)):.3f}%
α（生産者危険）: {self.app.controller.last_inputs.get('alpha', 5.0):.1f}%
β（消費者危険）: {self.app.controller.last_inputs.get('beta', 10.0):.1f}%
c値（許容不良数）: {self.app.controller.last_inputs.get('c_value', 0)}

{texts['review']}

{texts['best5']}

{self._get_adjustment_info()}

{self._get_guidance_info()}
"""
        
        try:
            filepath = filedialog.asksaveasfilename(
                title="結果を名前を付けて保存",
                defaultextension=".txt",
                filetypes=[("テキストファイル", "*.txt"), ("すべてのファイル", "*.*")],
                initialfile=f"検査結果_{self.app.controller.last_inputs['product_number']}.txt"
            )
            if not filepath: 
                return
            with open(filepath, 'w', encoding='utf-8') as f: 
                f.write(content)
            messagebox.showinfo("成功", f"結果を保存しました。\nパス: {filepath}")
        except Exception as e:
            sanitized_error = self.security_manager.sanitize_error_message(str(e))
            messagebox.showerror("エクスポート失敗", f"ファイルの保存中にエラーが発生しました: {sanitized_error}")
    
    def show_export_dialog(self):
        """エクスポート形式選択ダイアログ"""
        if not hasattr(self.app.controller, 'last_db_data') or not self.app.controller.last_db_data:
            messagebox.showinfo("エクスポート不可", "先に計算を実行してください。")
            return
        
        dialog = tk.Toplevel(self.app)
        dialog.title("レポート出力")
        dialog.geometry("450x400")
        dialog.minsize(450, 400)
        dialog.transient(self.app)
        dialog.grab_set()
        
        # メインフレーム
        main_frame = tk.Frame(dialog)
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        # タイトル
        title_label = ttk.Label(main_frame, text="出力形式を選択してください", 
                               font=('メイリオ', 12, 'bold'))
        title_label.pack(pady=(0, 20))
        
        # 出力形式選択
        format_frame = ttk.LabelFrame(main_frame, text="出力形式")
        format_frame.pack(fill='x', pady=(0, 20))
        
        format_var = tk.StringVar(value="text")
        
        # テキスト形式（常に利用可能）
        ttk.Radiobutton(format_frame, text="テキスト形式 (.txt)", 
                       variable=format_var, value="text").pack(anchor='w', padx=10, pady=5)
        
        # PDF形式
        if PDF_AVAILABLE:
            ttk.Radiobutton(format_frame, text="PDF形式 (.pdf)", 
                           variable=format_var, value="pdf").pack(anchor='w', padx=10, pady=5)
        else:
            ttk.Radiobutton(format_frame, text="PDF形式 (.pdf) - 未インストール", 
                           variable=format_var, value="pdf", state='disabled').pack(anchor='w', padx=10, pady=5)
        
        # Excel形式
        if EXCEL_AVAILABLE:
            ttk.Radiobutton(format_frame, text="Excel形式 (.xlsx)", 
                           variable=format_var, value="excel").pack(anchor='w', padx=10, pady=5)
        else:
            ttk.Radiobutton(format_frame, text="Excel形式 (.xlsx) - 未インストール", 
                           variable=format_var, value="excel", state='disabled').pack(anchor='w', padx=10, pady=5)
        
        # オプション
        options_frame = ttk.LabelFrame(main_frame, text="オプション")
        options_frame.pack(fill='x', pady=(0, 20))
        
        include_chart_var = tk.BooleanVar(value=True)
        if EXCEL_AVAILABLE:
            ttk.Checkbutton(options_frame, text="グラフを含める (Excel形式のみ)", 
                           variable=include_chart_var).pack(anchor='w', padx=10, pady=5)
        
        # ボタンフレーム
        button_frame = tk.Frame(main_frame)
        button_frame.pack(fill='x')
        
        def export_selected():
            format_type = format_var.get()
            
            if format_type == "text":
                dialog.destroy()
                self.export_results()
            elif format_type == "pdf" and PDF_AVAILABLE:
                dialog.destroy()
                self.export_pdf()
            elif format_type == "excel" and EXCEL_AVAILABLE:
                dialog.destroy()
                self.export_excel(include_chart_var.get())
            else:
                messagebox.showerror("エラー", "選択された形式は利用できません。")
        
        ttk.Button(button_frame, text="出力", command=export_selected).pack(side='right', padx=(5, 0))
        ttk.Button(button_frame, text="キャンセル", command=dialog.destroy).pack(side='right')
    
    def export_pdf(self):
        """PDF形式でのエクスポート"""
        if not PDF_AVAILABLE:
            messagebox.showerror("エラー", "PDF出力に必要なライブラリがインストールされていません。")
            return
        
        try:
            filepath = filedialog.asksaveasfilename(
                title="PDFファイルを保存",
                defaultextension=".pdf",
                filetypes=[("PDFファイル", "*.pdf")],
                initialfile=f"検査結果_{self.app.controller.last_inputs['product_number']}.pdf"
            )
            if not filepath:
                return
            
            self._generate_pdf_report(filepath)
            messagebox.showinfo("成功", f"PDFファイルを保存しました。\nパス: {filepath}")
            
        except Exception as e:
            sanitized_error = self.security_manager.sanitize_error_message(str(e))
            messagebox.showerror("エクスポート失敗", f"PDFファイルの保存中にエラーが発生しました: {sanitized_error}")
    
    def export_excel(self, include_chart=True):
        """Excel形式でのエクスポート"""
        if not EXCEL_AVAILABLE:
            messagebox.showerror("エラー", "Excel出力に必要なライブラリがインストールされていません。")
            return
        
        try:
            filepath = filedialog.asksaveasfilename(
                title="Excelファイルを保存",
                defaultextension=".xlsx",
                filetypes=[("Excelファイル", "*.xlsx")],
                initialfile=f"検査結果_{self.app.controller.last_inputs['product_number']}.xlsx"
            )
            if not filepath:
                return
            
            self._generate_excel_report(filepath, include_chart)
            messagebox.showinfo("成功", f"Excelファイルを保存しました。\nパス: {filepath}")
            
        except Exception as e:
            sanitized_error = self.security_manager.sanitize_error_message(str(e))
            messagebox.showerror("エクスポート失敗", f"Excelファイルの保存中にエラーが発生しました: {sanitized_error}")
    
    def _generate_pdf_report(self, filepath):
        """PDFレポートの生成（詳細版）"""
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # 日本語フォントの設定（既にインポート済み）
        
        try:
            # 日本語フォントの登録
            pdfmetrics.registerFont(UnicodeCIDFont('HeiseiKakuGo-W5'))
            font_name = 'HeiseiKakuGo-W5'
        except Exception as e:
            print(f"Font registration error (HeiseiKakuGo-W5): {e}")
            try:
                # 代替フォント
                pdfmetrics.registerFont(UnicodeCIDFont('HeiseiMin-W3'))
                font_name = 'HeiseiMin-W3'
            except Exception as e2:
                print(f"Font registration error (HeiseiMin-W3): {e2}")
                # フォールバック
                font_name = 'Helvetica'
        
        # カスタムスタイル
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.darkblue,
            fontName=font_name
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            textColor=colors.darkblue,
            fontName=font_name
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=10,
            fontName=font_name
        )
        
        # 結果テキストの生成（テキスト出力と同じ）
        texts = self.app.controller.ui_manager.generate_result_texts(
            self.app.controller.last_db_data, 
            self.app.controller.last_stats_results, 
            self.app.controller.last_inputs
        )
        sample_size_disp = self.app.controller.ui_manager.format_int(self.app.controller.last_stats_results['sample_size'])
        
        # タイトル
        title = Paragraph("AI SQC Sampler - 計算結果（AQL/LTPD設計）", title_style)
        story.append(title)
        story.append(Spacer(1, 20))
        
        # 基本情報
        story.append(Paragraph("基本情報", heading_style))
        basic_info = [
            ["品番", self.app.controller.last_inputs['product_number']],
            ["ロットサイズ", f"{self.app.controller.ui_manager.format_int(self.app.controller.last_inputs['lot_size'])}個"],
            ["不具合率", f"{self.app.controller.last_db_data['defect_rate']:.2f}%"],
            ["検査水準", self.app.controller.last_stats_results['level_text']],
            ["サンプルサイズ", f"{sample_size_disp}個"]
        ]
        
        basic_table = Table(basic_info, colWidths=[2*inch, 3*inch])
        basic_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (1, 0), (1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(basic_table)
        story.append(Spacer(1, 20))
        
        # AQL/LTPD設計パラメータ
        story.append(Paragraph("AQL/LTPD設計パラメータ", heading_style))
        stats_info = [
            ["AQL（合格品質水準）", f"{self.app.controller.last_stats_results.get('aql', self.app.controller.last_inputs.get('aql', 0.25)):.3f}%"],
            ["LTPD（不合格品質水準）", f"{self.app.controller.last_stats_results.get('ltpd', self.app.controller.last_inputs.get('ltpd', 1.0)):.3f}%"],
            ["α（生産者危険）", f"{self.app.controller.last_inputs.get('alpha', 5.0):.1f}%"],
            ["β（消費者危険）", f"{self.app.controller.last_inputs.get('beta', 10.0):.1f}%"],
            ["c値（許容不良数）", str(self.app.controller.last_inputs.get('c_value', 0))]
        ]
        
        stats_table = Table(stats_info, colWidths=[2.5*inch, 2.5*inch])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (1, 0), (1, -1), colors.lightcyan),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(stats_table)
        story.append(Spacer(1, 20))
        
        # レビュー情報
        if texts.get('review'):
            story.append(Paragraph("検査結果レビュー", heading_style))
            review_para = Paragraph(texts['review'], normal_style)
            story.append(review_para)
            story.append(Spacer(1, 15))
        
        # ベスト5情報
        if texts.get('best5'):
            story.append(Paragraph("推奨サンプルサイズ（上位5件）", heading_style))
            best5_para = Paragraph(texts['best5'], normal_style)
            story.append(best5_para)
            story.append(Spacer(1, 15))
        
        # 調整情報
        adjustment_info = self._get_adjustment_info()
        if adjustment_info:
            story.append(Paragraph("調整情報", heading_style))
            adjustment_para = Paragraph(adjustment_info, normal_style)
            story.append(adjustment_para)
            story.append(Spacer(1, 15))
        
        # ガイダンス情報
        guidance_info = self._get_guidance_info()
        if guidance_info:
            story.append(Paragraph("ガイダンス情報", heading_style))
            guidance_para = Paragraph(guidance_info, normal_style)
            story.append(guidance_para)
            story.append(Spacer(1, 15))
        
        # フッター
        footer = Paragraph(f"レポート生成日時: {datetime.now().strftime('%Y年%m月%d日 %H:%M:%S')}", 
                          normal_style)
        story.append(footer)
        
        doc.build(story)
    
    def _generate_excel_report(self, filepath, include_chart=True):
        """Excelレポートの生成（詳細版）"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "検査レポート"
        
        # 結果テキストの生成（テキスト出力と同じ）
        texts = self.app.controller.ui_manager.generate_result_texts(
            self.app.controller.last_db_data, 
            self.app.controller.last_stats_results, 
            self.app.controller.last_inputs
        )
        sample_size_disp = self.app.controller.ui_manager.format_int(self.app.controller.last_stats_results['sample_size'])
        
        # スタイル設定
        title_font = Font(name='メイリオ', size=16, bold=True, color='FFFFFF')
        header_font = Font(name='メイリオ', size=12, bold=True)
        normal_font = Font(name='メイリオ', size=10)
        
        title_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        
        # タイトル
        ws['A1'] = "AI SQC Sampler - 計算結果（AQL/LTPD設計）"
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:E1')
        
        # 基本情報
        row = 3
        ws[f'A{row}'] = "基本情報"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:E{row}')
        
        basic_data = [
            ["品番", self.app.controller.last_inputs['product_number']],
            ["ロットサイズ", f"{self.app.controller.ui_manager.format_int(self.app.controller.last_inputs['lot_size'])}個"],
            ["不具合率", f"{self.app.controller.last_db_data['defect_rate']:.2f}%"],
            ["検査水準", self.app.controller.last_stats_results['level_text']],
            ["サンプルサイズ", f"{sample_size_disp}個"]
        ]
        
        for i, (label, value) in enumerate(basic_data):
            row += 1
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = normal_font
            ws[f'B{row}'].font = normal_font
        
        # AQL/LTPD設計パラメータ
        row += 2
        ws[f'A{row}'] = "AQL/LTPD設計パラメータ"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:E{row}')
        
        stats_data = [
            ["AQL（合格品質水準）", f"{self.app.controller.last_stats_results.get('aql', self.app.controller.last_inputs.get('aql', 0.25)):.3f}%"],
            ["LTPD（不合格品質水準）", f"{self.app.controller.last_stats_results.get('ltpd', self.app.controller.last_inputs.get('ltpd', 1.0)):.3f}%"],
            ["α（生産者危険）", f"{self.app.controller.last_inputs.get('alpha', 5.0):.1f}%"],
            ["β（消費者危険）", f"{self.app.controller.last_inputs.get('beta', 10.0):.1f}%"],
            ["c値（許容不良数）", str(self.app.controller.last_inputs.get('c_value', 0))]
        ]
        
        for i, (label, value) in enumerate(stats_data):
            row += 1
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = normal_font
            ws[f'B{row}'].font = normal_font
        
        # レビュー情報
        if texts.get('review'):
            row += 2
            ws[f'A{row}'] = "検査結果レビュー"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws.merge_cells(f'A{row}:E{row}')
            
            row += 1
            ws[f'A{row}'] = texts['review']
            ws[f'A{row}'].font = normal_font
            ws.merge_cells(f'A{row}:E{row}')
            ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        
        # ベスト5情報
        if texts.get('best5'):
            row += 2
            ws[f'A{row}'] = "推奨サンプルサイズ（上位5件）"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws.merge_cells(f'A{row}:E{row}')
            
            row += 1
            ws[f'A{row}'] = texts['best5']
            ws[f'A{row}'].font = normal_font
            ws.merge_cells(f'A{row}:E{row}')
            ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        
        # 調整情報
        adjustment_info = self._get_adjustment_info()
        if adjustment_info:
            row += 2
            ws[f'A{row}'] = "調整情報"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws.merge_cells(f'A{row}:E{row}')
            
            row += 1
            ws[f'A{row}'] = adjustment_info
            ws[f'A{row}'].font = normal_font
            ws.merge_cells(f'A{row}:E{row}')
            ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        
        # ガイダンス情報
        guidance_info = self._get_guidance_info()
        if guidance_info:
            row += 2
            ws[f'A{row}'] = "ガイダンス情報"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws.merge_cells(f'A{row}:E{row}')
            
            row += 1
            ws[f'A{row}'] = guidance_info
            ws[f'A{row}'].font = normal_font
            ws.merge_cells(f'A{row}:E{row}')
            ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        
        # グラフの追加（オプション）
        if include_chart:
            try:
                row += 3
                chart_data = [
                    ["パラメータ", "値"],
                    ["AQL", self.app.controller.last_inputs.get('aql', 0.25)],
                    ["LTPD", self.app.controller.last_inputs.get('ltpd', 1.0)],
                    ["α", self.app.controller.last_inputs.get('alpha', 5.0)],
                    ["β", self.app.controller.last_inputs.get('beta', 10.0)]
                ]
                
                for i, (param, value) in enumerate(chart_data):
                    ws[f'G{row + i}'] = param
                    ws[f'H{row + i}'] = value
                
                # グラフの作成
                chart = BarChart()
                chart.title = "統計パラメータ"
                chart.x_axis.title = "パラメータ"
                chart.y_axis.title = "値 (%)"
                
                data = Reference(ws, min_col=8, min_row=row, max_row=row+4)
                cats = Reference(ws, min_col=7, min_row=row+1, max_row=row+4)
                chart.add_data(data, titles_from_data=False)
                chart.set_categories(cats)
                
                ws.add_chart(chart, f"G{row+6}")
            except Exception as e:
                print(f"Chart generation error: {e}")
                # グラフ生成に失敗した場合はスキップ
        
        # 列幅調整
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 30
        
        wb.save(filepath)
    
    def _get_adjustment_info(self):
        """調整情報の取得"""
        if 'adjustment_info' in self.app.controller.last_stats_results and self.app.controller.last_stats_results['adjustment_info']:
            return self.app.controller.last_stats_results['adjustment_info']
        return ""
    
    def _get_guidance_info(self):
        """ガイダンス情報の取得"""
        guidance_parts = []
        
        if 'guidance_message' in self.app.controller.last_stats_results and self.app.controller.last_stats_results['guidance_message']:
            guidance_parts.append(self.app.controller.last_stats_results['guidance_message'])
        
        if 'warning_message' in self.app.controller.last_stats_results and self.app.controller.last_stats_results['warning_message']:
            guidance_parts.append(f"【警告】\n{self.app.controller.last_stats_results['warning_message']}")
        
        return "\n\n".join(guidance_parts) if guidance_parts else ""